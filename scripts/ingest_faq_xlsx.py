#!/usr/bin/env python3
"""
ingest_faq_xlsx.py — Convert shiprocket_faq.xlsx → pillar_12_faq/ YAML files.

Usage:
    python3 scripts/ingest_faq_xlsx.py \
        --xlsx /path/to/shiprocket_faq.xlsx \
        --out  /path/to/knowledge_base/shiprocket/MultiChannel_API/pillar_12_faq

Each sheet tab → one subdirectory.
Each row (Q+A chunk) → one YAML file.

Quality gates applied:
  - Skip rows < 20 words after URL stripping
  - Strip embedded image URLs (amazonaws.com / S3 links)
  - Detect query_mode from question prefix
  - Assign trust_score: 0.7 (auto-generated FAQ content)
"""

import argparse
import os
import re
import sys
from pathlib import Path

import openpyxl
import yaml


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

MIN_WORD_COUNT = 20
TRUST_SCORE = 0.7

# Regex: strip image/screenshot URLs embedded in text
IMAGE_URL_RE = re.compile(
    r'https?://[^\s]*(?:amazonaws|s3|cdn|Screenshot)[^\s]*', re.IGNORECASE
)
# Generic bare URL cleanup (but keep API endpoint examples like /api/v1/...)
BARE_URL_RE = re.compile(r'https?://[^\s]{20,}')

# Map tab name → canonical domain label
TAB_DOMAIN_MAP = {
    "faq_orders":           "orders",
    "faq_returns":          "returns",
    "faq_miscellaneous":    "general",
    "faq_SRX":              "srx",
    "faq_checkout":         "checkout",
    "faq_brand_boost":      "brand_boost",
    "faq_engage":           "engage",
    "faq_fulfillment":      "fulfillment",
    "faq_omuni":            "omuni",
    "faq_promise":          "promise",
    "faq_AppStore":         "app_store",
    "faq_RTO":              "rto",
    "faq_secure":           "secure",
    "faq_sense":            "sense",
    "faq_instant_cod":      "instant_cod",
    "faq_trends":           "trends",
    "faq_quick":            "quick",
    "faq_zop":              "zop",
    "faq_credit_score":     "credit_score",
    "faq_buyer_protect":    "buyer_protect",
    "faq_profile":          "profile",
    "faq_API":              "api",
    "faq_weight":           "weight_dispute",
    "faq_buyer_experience": "buyer_experience",
    "faq_courier":          "courier",
    "faq_setup_manage":     "setup",
    "faq_tools":            "tools",
    "faq_settings":         "settings",
    "faq_ndr_deliveryBoost":"ndr",
    "faq_credit_line":      "credit_line",
    "faq_shipsure":         "shipsure",
    "faq_home":             "home",
    "faq_dashboard":        "dashboard",
    "faq_finance":          "finance",
    "faq_navigation_bar":   "navigation",
    "faq_business_loan":    "business_loan",
    "faq_cargo":            "cargo",
    "faq_revprotect":       "revprotect",
    "faq_self_serve_labels":"labels",
    "faq_blogs":            "general",
}


def detect_query_mode(question: str) -> str:
    """Infer query_mode from the question prefix."""
    q = question.lower().strip()
    if any(q.startswith(p) for p in ("why ", "what causes", "reason for", "why is", "why does")):
        return "diagnose"
    if any(q.startswith(p) for p in ("where ", "how to access", "how to find", "how to navigate",
                                      "how to view", "how to open", "how to go")):
        return "navigate"
    if any(q.startswith(p) for p in ("what is ", "what are ", "what does ", "define ", "explain ",
                                      "tell me about", "describe ")):
        return "explain"
    if any(q.startswith(p) for p in ("how to ", "how do i", "how can i", "steps to", "can i ",
                                      "is it possible", "how to use")):
        return "lookup"
    return "lookup"


def extract_keywords(question: str, answer: str, domain: str) -> list:
    """Extract simple keyword hints from the question."""
    words = re.findall(r'\b[A-Za-z]{4,}\b', question)
    # Deduplicate, lowercase, remove stopwords
    stopwords = {"what", "does", "have", "with", "from", "that", "this", "your",
                 "will", "when", "which", "where", "there", "their", "shiprocket",
                 "using", "into", "been", "able", "also", "just", "more"}
    keywords = list(dict.fromkeys(
        w.lower() for w in words if w.lower() not in stopwords
    ))[:8]
    if domain not in keywords:
        keywords.insert(0, domain)
    return keywords


def clean_text(text: str) -> str:
    """Strip image URLs, normalize whitespace."""
    text = IMAGE_URL_RE.sub('', text)
    # Strip lines that are ONLY a URL (after image strip, may leave blank lines)
    lines = []
    for line in text.split('\n'):
        stripped = line.strip()
        if stripped and not re.match(r'^https?://\S+$', stripped):
            lines.append(line)
    return '\n'.join(lines).strip()


def split_qa(chunk: str) -> tuple[str, str]:
    """Split chunk_content into question (first line) and answer (rest)."""
    lines = chunk.strip().split('\n')
    question = lines[0].strip().rstrip('?').strip() + '?'
    answer = '\n'.join(lines[1:]).strip()
    return question, answer


def chunk_to_yaml(
    tab_name: str,
    domain: str,
    idx: int,
    question: str,
    answer: str,
    word_count: int,
) -> dict:
    query_mode = detect_query_mode(question)
    keywords = extract_keywords(question, answer, domain)

    # Build canonical_summary: first 2 sentences of answer, max 150 chars
    answer_sentences = re.split(r'(?<=[.!?])\s+', answer.replace('\n', ' '))
    summary = ' '.join(answer_sentences[:2])[:200].strip()
    if len(summary) < 30:
        summary = answer[:200].strip()

    return {
        "_tier": "high",
        "_source": "shiprocket_faq.xlsx",
        "_tab": tab_name,
        "_word_count": word_count,
        "entity_type": "faq_chunk",
        "entity_id": f"faq:{domain}:{idx:04d}",
        "domain": domain,
        "query_mode": query_mode,
        "trust_score": TRUST_SCORE,
        "question": question,
        "answer": answer,
        "canonical_summary": summary,
        "keywords": keywords,
        "related_domains": [domain],
    }


def process_xlsx(xlsx_path: str, out_dir: str) -> dict:
    """Main processing loop. Returns stats dict."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)

    stats = {
        "total_rows": 0,
        "written": 0,
        "skipped_short": 0,
        "skipped_empty": 0,
        "tabs_processed": 0,
        "by_tab": {},
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        domain = TAB_DOMAIN_MAP.get(sheet_name, sheet_name.replace("faq_", ""))

        tab_dir = out_path / sheet_name
        tab_dir.mkdir(exist_ok=True)

        tab_written = 0
        tab_skipped = 0
        idx = 0

        for row_num, row in enumerate(ws.iter_rows(values_only=True)):
            if row_num == 0:
                # Skip header row
                continue
            if not row or row[0] is None:
                stats["skipped_empty"] += 1
                continue

            raw = str(row[0]).strip()
            stats["total_rows"] += 1

            # Clean: strip image URLs
            cleaned = clean_text(raw)
            word_count = len(cleaned.split())

            if word_count < MIN_WORD_COUNT:
                stats["skipped_short"] += 1
                tab_skipped += 1
                continue

            question, answer = split_qa(cleaned)

            # Skip if answer is empty after split
            if len(answer.strip()) < 10:
                stats["skipped_short"] += 1
                tab_skipped += 1
                continue

            idx += 1
            doc = chunk_to_yaml(sheet_name, domain, idx, question, answer, word_count)

            yaml_path = tab_dir / f"faq_{idx:04d}.yaml"
            with open(yaml_path, "w", encoding="utf-8") as f:
                yaml.dump(doc, f, allow_unicode=True, default_flow_style=False,
                          sort_keys=False, width=120)

            tab_written += 1
            stats["written"] += 1

        stats["by_tab"][sheet_name] = {"written": tab_written, "skipped": tab_skipped}
        stats["tabs_processed"] += 1

    wb.close()
    return stats


def main():
    parser = argparse.ArgumentParser(description="Convert shiprocket_faq.xlsx to pillar_12_faq YAMLs")
    parser.add_argument("--xlsx", required=True, help="Path to shiprocket_faq.xlsx")
    parser.add_argument("--out",  required=True, help="Output directory (pillar_12_faq/)")
    args = parser.parse_args()

    if not os.path.exists(args.xlsx):
        print(f"ERROR: file not found: {args.xlsx}", file=sys.stderr)
        sys.exit(1)

    print(f"Processing: {args.xlsx}")
    print(f"Output dir: {args.out}")
    print()

    stats = process_xlsx(args.xlsx, args.out)

    print(f"Results:")
    print(f"  Tabs processed : {stats['tabs_processed']}")
    print(f"  Total rows     : {stats['total_rows']}")
    print(f"  Written        : {stats['written']}")
    print(f"  Skipped short  : {stats['skipped_short']}")
    print(f"  Skipped empty  : {stats['skipped_empty']}")
    print()
    print(f"Per-tab breakdown:")
    for tab, s in sorted(stats["by_tab"].items(), key=lambda x: -x[1]["written"]):
        print(f"  {tab:<30}  written={s['written']:>4}  skipped={s['skipped']:>3}")


if __name__ == "__main__":
    main()
